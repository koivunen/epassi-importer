
from base64 import b64encode
import os,requests,sys,time,json
from dotenv import load_dotenv

import pprint
from requests.auth import HTTPDigestAuth
from io import BytesIO
from datetime import datetime, time
from openpyxl import load_workbook
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import psycopg2
from psycopg2.extras import execute_values

def extract_timestamps(sheet):
    timestamps = []
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
        timestamp_str = row[0]
        try:
            timestamp = datetime.strptime(timestamp_str, "%d.%m.%Y %H:%M ")
            timestamps.append(timestamp)
        except (TypeError, ValueError):
            continue
    return timestamps


# chatgpt chaos
# TODO: replace with below or remove
def count_rows_in_window_per_day(xlsx_bytes: bytes, sheet_name: str | None = None,
                                 start: time = time(10, 20), end: time = time(13, 40)) -> dict:
    wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    time_col_idx = 0
    if header and isinstance(header, (list, tuple)):
        lowered = [str(v).strip().lower() if v is not None else "" for v in header]
        if "aika" in lowered:
            time_col_idx = lowered.index("aika")
        elif "time" in lowered:
            time_col_idx = lowered.index("time")
        elif "timestamp" in lowered:
            time_col_idx = lowered.index("timestamp")
    counts: dict = {}
    for row in rows:
        if not row:
            continue
        v = row[time_col_idx] if time_col_idx < len(row) else None
        if v is None:
            continue
        if isinstance(v, datetime):
            dt = v
        else:
            s = str(v).strip()
            if not s or s.lower().startswith("restaurant"):
                continue
            dt = None
            for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S",
                        "%m/%d/%Y %H:%M", "%m/%d/%Y %H:%M:%S"):
                try:
                    dt = datetime.strptime(s, fmt)
                    break
                except ValueError:
                    continue
            if dt is None:
                continue
        tt = dt.time()
        if tt < start or tt > end:
            continue
        d = dt.date()
        counts[d] = counts.get(d, 0) + 1
    return counts


from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import psycopg2
from psycopg2.extras import execute_values
from io import BytesIO

def extract_timestamps(sheet):
    timestamps = []
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
        timestamp_str = row[0]
        try:
            timestamp = datetime.strptime(timestamp_str, "%d.%m.%Y %H:%M ")
            timestamps.append(timestamp)
        except (TypeError, ValueError):
            continue
    return timestamps

def bucket_timestamps(timestamps) -> dict[datetime, int]:
    timestamps_df = pd.Series(timestamps)
    timestamps_df = pd.to_datetime(timestamps_df)
    bucketed_timestamps = timestamps_df.dt.floor('min')  
    
    # Bucket to 5-minute intervals manually
    bucketed_timestamps = bucketed_timestamps - pd.to_timedelta(bucketed_timestamps.dt.minute % 5, unit='m')
    
    ret = bucketed_timestamps.value_counts().sort_index()
    return ret.to_dict()


if __name__ == "__main__":
    with open('test.xlsx', 'rb') as file:
        file_content = file.read()
    
    byte_stream = BytesIO(file_content)
    workbook = load_workbook(byte_stream)
    sheet = workbook['Sheet1']
    timestamps = extract_timestamps(sheet)
    bucketed_timestamps = bucket_timestamps(timestamps)
    print("Bucketed Timestamps:")
    print(bucketed_timestamps)
    
    #store_to_postgresql(bucketed_timestamps, db_params)


    with open("test.xlsx", "rb") as f:
        xlsx_bytes = f.read()
    counts = count_rows_in_window_per_day(xlsx_bytes)
    for d, c in counts.items():
        print(f"{d}: {c}")